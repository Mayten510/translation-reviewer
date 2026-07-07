#!/usr/bin/env python3
"""
Multi-format source-target alignment for translation review.
- Excel: cell-by-cell alignment with omission/addition detection
- PDF/DOCX/PPTX/MD/HTML: section-level extraction from both files

Usage:
    python compare.py <source> <target>    # auto-detect format
    python compare.py --text <source.txt> <target.txt>   # plain text alignment
"""

import json
import re
import sys
from pathlib import Path
from collections import defaultdict


def has_chinese(text):
    if not text or not isinstance(text, str):
        return False
    return bool(re.search(r'[一-鿿]', text))


# ─── Import extract.py's parsers ───
# We reuse extract logic internally for non-Excel formats

EXTRACT_SCRIPT = Path(__file__).parent / "extract.py"


def run_extract(path, is_source=False):
    """Run extract.py and return parsed sections."""
    import subprocess
    args = [sys.executable, str(EXTRACT_SCRIPT), str(path)]
    if is_source:
        args.append("--source")
    result = subprocess.run(args, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"Extract error: {result.stderr}", file=sys.stderr)
        return None
    return json.loads(result.stdout)


# ─── Excel comparison (cell-by-cell) ───

def extract_excel_cells(ws):
    cells = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    cells[cell.coordinate] = val
    return cells


def compare_excel(source_path, target_path):
    import openpyxl
    wb_src = openpyxl.load_workbook(source_path, data_only=True)
    wb_tgt = openpyxl.load_workbook(target_path, data_only=True)

    result = {
        "mode": "excel",
        "source_file": str(source_path),
        "target_file": str(target_path),
        "sheets": [],
        "pairs": [],
        "findings": [],
        "consistency": {},
        "stats": defaultdict(int),
    }

    all_pairs = []
    total_stats = defaultdict(int)

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]

        # Find matching target sheet
        ws_tgt = None
        if sheet_name in wb_tgt.sheetnames:
            ws_tgt = wb_tgt[sheet_name]
        else:
            for tgt_name in wb_tgt.sheetnames:
                base = tgt_name.rstrip('_EN_JA_KO_DE_FR_ES')
                if sheet_name == base or tgt_name.startswith(sheet_name):
                    ws_tgt = wb_tgt[tgt_name]
                    break
            if ws_tgt is None:
                src_idx = wb_src.sheetnames.index(sheet_name)
                if src_idx < len(wb_tgt.sheetnames):
                    ws_tgt = wb_tgt[wb_tgt.sheetnames[src_idx]]

        if ws_tgt is None:
            result["findings"].append({
                "dimension": "漏翻", "severity": "P1",
                "location": f"Sheet: {sheet_name}",
                "description": f"No matching sheet in target for '{sheet_name}'",
            })
            continue

        src_cells = extract_excel_cells(ws_src)
        tgt_cells = extract_excel_cells(ws_tgt)
        all_coords = set(src_cells.keys()) | set(tgt_cells.keys())

        sheet_pairs = []
        sheet_stats = defaultdict(int)

        def coord_key(c):
            m = re.match(r'([A-Z]+)(\d+)', c)
            if not m: return (0, 0)
            cn = 0
            for ch in m.group(1): cn = cn * 26 + (ord(ch) - ord('A') + 1)
            return (cn, int(m.group(2)))

        for coord in sorted(all_coords, key=coord_key):
            sv = src_cells.get(coord)
            tv = tgt_cells.get(coord)
            match_type = "translated" if (sv and tv and has_chinese(sv)) else \
                         "missing" if (sv and has_chinese(sv) and not tv) else \
                         "addition" if (not sv and tv) else \
                         "unchanged" if (sv and tv and sv == tv) else "modified"
            pair = {"coordinate": coord, "source": sv, "target": tv, "match": match_type}
            sheet_pairs.append(pair)
            sheet_stats[match_type] += 1
            total_stats[match_type] += 1

            if match_type == "missing":
                result["findings"].append({
                    "dimension": "漏译", "severity": "P1",
                    "location": f"{sheet_name}!{coord}",
                    "description": f"Source '{sv[:50]}' not translated",
                })
            elif match_type == "addition" and tv and len(tv) > 3:
                result["findings"].append({
                    "dimension": "增译", "severity": "P2",
                    "location": f"{sheet_name}!{coord}",
                    "description": f"Added '{tv[:50]}' with no source",
                })

        all_pairs.extend(sheet_pairs)
        result["sheets"].append({
            "source_name": sheet_name,
            "target_name": ws_tgt.title,
            "pairs": sheet_pairs,
            "stats": dict(sheet_stats),
        })

    wb_src.close()
    wb_tgt.close()

    # Consistency check
    groups = defaultdict(list)
    for p in all_pairs:
        if p["source"] and has_chinese(p["source"]):
            groups[p["source"]].append({"coordinate": p["coordinate"], "target": p["target"]})
    consistency_issues = []
    for src_text, occs in groups.items():
        if len(occs) > 1:
            targets = [o["target"] for o in occs if o["target"]]
            ut = list(set(targets))
            if len(ut) > 1:
                consistency_issues.append({"source": src_text, "unique_translations": ut, "occurrences": occs})
                result["findings"].append({
                    "dimension": "一致性", "severity": "P1" if len(ut) > 2 else "P2",
                    "location": ", ".join(o["coordinate"] for o in occs[:5]),
                    "description": f"'{src_text[:30]}' translated {len(ut)} ways: {ut}",
                })
    result["consistency"] = {
        "total_unique": len(groups),
        "issues": consistency_issues,
    }
    result["stats"] = dict(total_stats)
    result["total_pairs"] = len(all_pairs)
    result["total_findings"] = len(result["findings"])

    return result


# ─── Generic text comparison (for PDF/DOCX/PPTX/MD/HTML) ───

def compare_generic(source_path, target_path):
    src = run_extract(source_path, is_source=True)
    tgt = run_extract(target_path, is_source=False)

    result = {
        "mode": "generic",
        "source_file": str(source_path),
        "target_file": str(target_path),
        "source_format": src["format"] if src else "unknown",
        "target_format": tgt["format"] if tgt else "unknown",
        "source_sections": src["sections"][:200] if src else [],
        "target_sections": tgt["sections"][:200] if tgt else [],
        "source_count": src["total_sections"] if src else 0,
        "target_count": tgt["total_sections"] if tgt else 0,
        "automated_findings": tgt["automated_findings"] if tgt else [],
        "automated_count": tgt["automated_count"] if tgt else 0,
        "review_texts": tgt["review_texts"] if tgt else [],
    }

    # Flag size mismatches
    if src and tgt:
        ratio = tgt["total_sections"] / max(src["total_sections"], 1)
        if ratio > 1.5:
            result["size_warning"] = f"Target is {ratio:.1f}x larger than source — possible additions or expanded version"
        elif ratio < 0.5:
            result["size_warning"] = f"Target is {ratio:.1f}x smaller than source — possible omissions"

    result["total_findings"] = len(result["automated_findings"])
    return result


# ─── Main ───

def main():
    if len(sys.argv) < 3:
        print("Usage: python compare.py <source> <target>", file=sys.stderr)
        print("       python compare.py --text <source.txt> <target.txt>", file=sys.stderr)
        sys.exit(1)

    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    if len(args) < 2:
        print("Error: need source and target files", file=sys.stderr)
        sys.exit(1)

    source_path = Path(args[0])
    target_path = Path(args[1])

    for p in [source_path, target_path]:
        if not p.exists():
            print(f"Error: file not found: {p}", file=sys.stderr)
            sys.exit(1)

    ext = source_path.suffix.lower()
    if ext in ('.xlsx', '.xls'):
        result = compare_excel(str(source_path), str(target_path))
    else:
        result = compare_generic(str(source_path), str(target_path))

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
