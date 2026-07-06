#!/usr/bin/env python3
"""
Export translation review findings to a formatted Excel report.

Input: JSON array of findings + optional summary
Output: .xlsx with two sheets — "审查结果" (detailed) and "审查总结" (summary)

Usage:
    python export_review.py <findings.json> <output.xlsx>
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Severity color scheme
SEVERITY_FILLS = {
    "P0": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "P1": PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid"),
    "P2": PatternFill(start_color="FFDD44", end_color="FFDD44", fill_type="solid"),
    "P3": PatternFill(start_color="88CCEE", end_color="88CCEE", fill_type="solid"),
}

SEVERITY_FONTS = {
    "P0": Font(bold=True, color="FFFFFF", size=10),
    "P1": Font(bold=True, color="FFFFFF", size=10),
    "P2": Font(bold=True, size=10),
    "P3": Font(size=10),
}

DIMENSIONS = [
    "漏翻", "错译", "一致性", "术语错误", "拼写错误",
    "语法错误", "敏感词", "技术原理", "翻译规范", "地道性",
]


def make_styles():
    """Return common style objects."""
    return {
        "header_font": Font(bold=True, size=11, color="FFFFFF"),
        "header_fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "cell_align": Alignment(vertical="center", wrap_text=True),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "cell_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "total_fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "bold_font": Font(bold=True, size=11),
        "normal_font": Font(size=11),
    }


def build_detail_sheet(ws, findings, st):
    """Build the detailed findings sheet."""
    headers = ["#", "问题位置", "问题描述", "严重程度", "问题类别"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = st["header_font"]
        c.fill = st["header_fill"]
        c.alignment = st["header_align"]
        c.border = st["cell_border"]

    for row_idx, finding in enumerate(findings, 2):
        severity = finding.get("severity", "P3")
        row_data = [
            row_idx - 1,
            finding.get("location", ""),
            finding.get("description", ""),
            severity,
            finding.get("dimension", ""),
        ]

        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = st["cell_align"]
            c.border = st["cell_border"]

            if col_idx == 4:
                c.fill = SEVERITY_FILLS.get(severity, PatternFill())
                c.font = SEVERITY_FONTS.get(severity, Font(size=10))
                c.alignment = st["center_align"]
            elif col_idx == 1:
                c.alignment = st["center_align"]

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16

    # Freeze header
    ws.freeze_panes = "A2"


def build_summary_sheet(ws, findings, st, assessment=""):
    """Build the summary sheet with counts per dimension."""
    headers = ["维度", "P0", "P1", "P2", "P3", "合计"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = st["header_font"]
        c.fill = st["header_fill"]
        c.alignment = st["header_align"]
        c.border = st["cell_border"]

    # Count findings per dimension and severity
    counts = {dim: {"P0": 0, "P1": 0, "P2": 0, "P3": 0} for dim in DIMENSIONS}
    for f in findings:
        dim = f.get("dimension", "")
        sev = f.get("severity", "P3")
        if dim in counts:
            counts[dim][sev] += 1

    grand = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    for row_idx, dim in enumerate(DIMENSIONS, 2):
        dim_counts = counts[dim]
        total = sum(dim_counts.values())
        row_data = [dim, dim_counts["P0"], dim_counts["P1"], dim_counts["P2"], dim_counts["P3"], total]
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = st["cell_border"]
            c.alignment = st["center_align"]
            if col_idx == 1:
                c.font = Font(bold=True, size=10)
        for sev in ["P0", "P1", "P2", "P3"]:
            grand[sev] += dim_counts[sev]

    # Grand total row
    total_row = len(DIMENSIONS) + 2
    grand_total = sum(grand.values())
    for col_idx, val in enumerate(["总计", grand["P0"], grand["P1"], grand["P2"], grand["P3"], grand_total], 1):
        c = ws.cell(row=total_row, column=col_idx, value=val)
        c.font = st["bold_font"]
        c.fill = st["total_fill"]
        c.border = st["cell_border"]
        c.alignment = st["center_align"]

    # Assessment row
    if assessment:
        assess_row = total_row + 2
        ws.merge_cells(start_row=assess_row, start_column=1, end_row=assess_row, end_column=6)
        c = ws.cell(row=assess_row, column=1, value=assessment)
        c.font = st["normal_font"]
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = st["cell_border"]
        ws.row_dimensions[assess_row].height = 40

    # Column widths
    ws.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 8

    ws.freeze_panes = "A2"


def main():
    if len(sys.argv) < 3:
        print("Usage: python export_review.py <findings.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    findings_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    with open(findings_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Support both flat list and {findings: [...], assessment: "..."}
    if isinstance(data, list):
        findings = data
        assessment = ""
    else:
        findings = data.get("findings", [])
        assessment = data.get("assessment", "")

    st = make_styles()
    wb = openpyxl.Workbook()

    # Detail sheet
    ws1 = wb.active
    ws1.title = "审查结果"
    build_detail_sheet(ws1, findings, st)

    # Summary sheet
    ws2 = wb.create_sheet("审查总结")
    build_summary_sheet(ws2, findings, st, assessment)

    wb.save(str(output_path))
    print(json.dumps({
        "status": "success",
        "output_file": str(output_path.absolute()),
        "findings_count": len(findings),
        "sheets": wb.sheetnames,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
