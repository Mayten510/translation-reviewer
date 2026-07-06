#!/usr/bin/env python3
"""
Automated first-pass scan of translated Excel files.

Checks:
- 漏翻: Remaining Chinese characters (CJK)
- 拼写错误: Basic spellcheck for non-CJK languages
- Extracts all unique texts grouped by sheet for manual review

Outputs a JSON structure ready for the manual review phase.
"""

import json
import re
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---- Language detection patterns ----

LANG_PATTERNS = {
    "EN": re.compile(r'[a-zA-Z]{3,}'),  # 3+ Latin letters → likely English text
    "JA": re.compile(r'[ぁ-んァ-ン]'),  # Kana → Japanese
    "KO": re.compile(r'[가-힣]'),  # Hangul → Korean
    "DE": re.compile(r'[äöüßÄÖÜ]'),  # German special chars
    "FR": re.compile(r'[éèêëàâùûîïôçÉÈÊËÀÂÙÛÎÏÔÇ]'),  # French accents
    "ES": re.compile(r'[áéíóúñÁÉÍÓÚÑüÜ]'),  # Spanish special chars
}

LANG_NAMES = {
    "EN": "English", "JA": "Japanese", "KO": "Korean",
    "DE": "German", "FR": "French", "ES": "Spanish",
}


def has_chinese(text):
    """Check if text contains Chinese characters."""
    if not text or not isinstance(text, str):
        return False
    return bool(re.search(r'[一-鿿]', text))


def detect_language(text):
    """
    Detect the most likely language of a text.
    Returns language code or 'unknown'.
    """
    if not text or not isinstance(text, str) or not text.strip():
        return "unknown"

    scores = {}
    for lang, pattern in LANG_PATTERNS.items():
        matches = pattern.findall(text)
        if matches:
            scores[lang] = len(matches)

    if not scores:
        # Pure ASCII/punctuation/numbers — default to EN for review
        return "EN"

    return max(scores, key=scores.get)


def is_sensitive(text):
    """
    Check for potentially sensitive terms.
    Returns list of flagged terms.
    """
    sensitive_patterns = [
        (r'\bTaiwan\b(?!\s+Strait)', 'P0', '可能涉及台湾表述'),
        (r'\bTibet\b', 'P0', '可能涉及西藏表述'),
        (r'\b(Communist|CCP|Mao|Tiananmen)\b', 'P0', '敏感政治词汇'),
        (r'\b(drug|narcotic|cocaine|heroin)\b', 'P1', '敏感词汇'),
        (r'\b(hacker?|crack(er|ing)|exploit|backdoor)\b', 'P2', '安全敏感词'),
        (r'\b(kill|destroy|attack|bomb)\b', 'P1', '暴力相关词汇'),
        (r'\b(fraud|scam|illegal|criminal)\b', 'P1', '法律敏感词'),
    ]

    flagged = []
    for pattern, severity, desc in sensitive_patterns:
        matches = list(re.finditer(pattern, text, re.IGNORECASE))
        for m in matches:
            # Check context — some terms are fine in tech context
            if desc == '安全敏感词' and 'security' in text.lower():
                continue  # Technical security context is OK
            flagged.append({
                "term": m.group(),
                "severity": severity,
                "description": desc,
            })

    return flagged


def extract_sheet_data(ws):
    """Extract all translatable content from a sheet."""
    cells = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                text = str(cell.value).strip()
                lang = detect_language(text)

                cells.append({
                    "coordinate": cell.coordinate,
                    "value": text,
                    "detected_lang": lang,
                    "has_chinese": has_chinese(text),
                    "char_count": len(text),
                })

    # Group by language for review
    by_lang = defaultdict(list)
    for c in cells:
        by_lang[c["detected_lang"]].append(c)

    return {
        "name": ws.title,
        "total_cells": len(cells),
        "by_language": {k: v for k, v in by_lang.items()},
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python scan.py <translated.xlsx>", file=sys.stderr)
        sys.exit(1)

    source_path = Path(sys.argv[1])
    if not source_path.exists():
        print(f"Error: file not found: {source_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(source_path, data_only=True)

    result = {
        "source_file": str(source_path.absolute()),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "sheets": [],
        "automated_findings": [],
    }

    all_texts = []  # All unique texts for consistency checking

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = extract_sheet_data(ws)

        # ---- Automated checks ----

        # 1. 漏翻 check
        for lang_group in sheet_data["by_language"].values():
            for cell in lang_group:
                if cell["has_chinese"]:
                    result["automated_findings"].append({
                        "dimension": "漏翻",
                        "severity": "P1",
                        "location": f"{sheet_name}!{cell['coordinate']}",
                        "description": f"Cell contains untranslated Chinese: '{cell['value'][:60]}'",
                    })

                # 7. 敏感词 check
                sensitive = is_sensitive(cell["value"])
                for s in sensitive:
                    result["automated_findings"].append({
                        "dimension": "敏感词",
                        "severity": s["severity"],
                        "location": f"{sheet_name}!{cell['coordinate']}",
                        "description": f"{s['description']}: '{s['term']}' in '{cell['value'][:60]}'",
                    })

        # 5. 拼写检查 — skipped in automated pass, done by Claude in manual review
        # Spellcheck is too noisy with a basic word list; pyspellchecker is optional.
        # The scan output provides sample_texts for Claude to review manually.

        # Collect unique texts for consistency review
        for lang_group in sheet_data["by_language"].values():
            for cell in lang_group:
                all_texts.append({
                    "sheet": sheet_name,
                    "coordinate": cell["coordinate"],
                    "value": cell["value"],
                    "lang": cell["detected_lang"],
                })

        # Don't send all cell data — just summary for the JSON output
        sheet_summary = {
            "name": sheet_data["name"],
            "total_cells": sheet_data["total_cells"],
            "languages": list(sheet_data["by_language"].keys()),
            "sample_texts": [],
        }

        # Pick representative samples per language for manual review
        for lang, cells in sheet_data["by_language"].items():
            # Take up to 25 samples per language per sheet
            for c in cells[:25]:
                sheet_summary["sample_texts"].append({
                    "coordinate": c["coordinate"],
                    "value": c["value"],
                    "lang": lang,
                })

        result["sheets"].append(sheet_summary)

    wb.close()

    # Summary
    result["total_automated_findings"] = len(result["automated_findings"])
    result["findings_by_dimension"] = defaultdict(int)
    for f in result["automated_findings"]:
        result["findings_by_dimension"][f["dimension"]] += 1
    result["findings_by_dimension"] = dict(result["findings_by_dimension"])

    # Group findings by dimension
    by_dim = defaultdict(list)
    for f in result["automated_findings"]:
        by_dim[f["dimension"]].append(f)
    result["automated_findings_by_dimension"] = {k: v for k, v in by_dim.items()}

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
